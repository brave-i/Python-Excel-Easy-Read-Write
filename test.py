import openpyxl

book = None

def Cert_Updates(file_Name, sheet_Name):

	# open worksheet
	print ("-------   opening excel file... it will take a few sec  ---------")

	book = openpyxl.load_workbook(file_Name)

	# open sheet
	sheet = book.get_sheet_by_name(sheet_Name)
	print(sheet.title)

	nIndex = 2

	Cert_Name = "Cert"
	Display_Name = "DisplayName"
	Email_Addr = "Email"

	changestate = False
	# cell = sheet.cell(row=28, column=3)
	# print (cell.value)



	while (Display_Name!=None):

		# getting cert name
		
		Cert_Cell = sheet.cell(row = nIndex, column=3)
		Cert_Name = Cert_Cell.value

		if Cert_Name == "Steno":
			Cert_Cell.value = "Voice"
			changestate = True

		#getting display Name

		Display_Cell = sheet.cell(row = nIndex, column=4)
		Display_Name = Display_Cell.value

		if Display_Name == None:
			continue

		#getting Email Address

		Email_Cell = sheet.cell(row = nIndex, column=8)
		Email_Addr = Email_Cell.value

		if changestate == True:
			print (nIndex)
			print (Email_Addr)
			changestate = False


		nIndex = nIndex + 1

	print ("--------------------  saving now...  -----------------------")

	book.save('export.xlsx')
	print ("-------------------- saving done, thank you! ---------------------")


try:
	
	fileName = 'example.xlsx'
	sheetName = "Local Overflow"

	Cert_Updates(fileName, sheetName)
	
except Exception as e:
	print ("there is no excel file, please confirm again!")